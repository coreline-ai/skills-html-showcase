#!/usr/bin/env python3
"""
X(트위터) AI 트렌드 수집 결과를 엑셀 + HTML 대시보드로 출력한다.

핵심 동작:
- 같은 이름의 엑셀이 이미 있으면 '덮어쓰지 않고' 기존 데이터에 새 데이터를 이어붙인다(append).
- 트윗 URL을 기준으로 중복을 제거한다(같은 URL은 새 데이터로 갱신).
- 합쳐진 전체 데이터로 엑셀과 HTML 대시보드를 다시 렌더링한다.

사용법:
    python build_outputs.py --input records.json --outdir <폴더> [--basename AI_트렌드]

records.json 형식 (리스트):
[
  {
    "cat": "신규 모델·제품 출시",     # 카테고리(아래 CATEGORIES 중 하나 권장)
    "author": "Anthropic (공식)",      # 표시 이름
    "handle": "@claudeai",             # @핸들
    "date": "2026-05-29",              # YYYY-MM-DD
    "summary": "핵심만 1~2문장으로 요약", # 한국어 요약
    "url": "https://x.com/claudeai/status/123",  # 트윗 원문 URL (중복 판단 기준)
    "views": 210468,                    # 조회수(정수, 모르면 0)
    "likes": 1133                       # 좋아요(정수, 모르면 0)
  }
]
"""
import argparse, json, os, sys, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "맑은 고딕"
NAVY = "1F2A44"; ACCENT = "2D6CDF"; LIGHT = "EAF0FB"; GRAYB = "D9DEE8"
CATEGORIES = ["신규 모델·제품 출시", "연구·논문 동향", "업계·투자 뉴스", "실무 팁·도구"]
CAT_COLOR = {"신규 모델·제품 출시": "#2D6CDF", "연구·논문 동향": "#7C3AED",
             "업계·투자 뉴스": "#34d399", "실무 팁·도구": "#fbbf24"}
HEADERS = ["번호", "카테고리", "작성자", "핸들", "날짜", "핵심 요약", "링크", "조회수", "좋아요"]


def norm(rec):
    """레코드를 표준 dict로 정규화."""
    return {
        "cat": str(rec.get("cat", "기타")).strip(),
        "author": str(rec.get("author", "")).strip(),
        "handle": str(rec.get("handle", "")).strip(),
        "date": str(rec.get("date", "")).strip(),
        "summary": str(rec.get("summary", "")).strip(),
        "url": str(rec.get("url", "")).strip(),
        "views": int(rec.get("views", 0) or 0),
        "likes": int(rec.get("likes", 0) or 0),
    }


def read_existing_xlsx(path):
    """기존 엑셀에서 데이터 행을 복원한다. 실패 시 빈 리스트."""
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path)
        ws = wb.active
        # 헤더('번호') 행을 찾는다
        hdr = None
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value).strip() == "번호":
                hdr = r
                break
        if hdr is None:
            return []
        out = []
        for r in range(hdr + 1, ws.max_row + 1):
            no = ws.cell(r, 1).value
            if not isinstance(no, int):  # 합계행/빈행에서 종료
                break
            link_cell = ws.cell(r, 7)
            url = link_cell.hyperlink.target if link_cell.hyperlink else ""
            out.append(norm({
                "cat": ws.cell(r, 2).value, "author": ws.cell(r, 3).value,
                "handle": ws.cell(r, 4).value, "date": ws.cell(r, 5).value,
                "summary": ws.cell(r, 6).value, "url": url,
                "views": ws.cell(r, 8).value or 0, "likes": ws.cell(r, 9).value or 0,
            }))
        return out
    except Exception as e:
        print(f"[warn] 기존 엑셀을 읽지 못해 새로 만듭니다: {e}", file=sys.stderr)
        return []


def merge(existing, new):
    """URL 기준 병합. 같은 URL은 새 데이터로 갱신. URL 없으면 그대로 추가."""
    by_url = {}
    order = []
    for rec in existing + new:
        key = rec["url"] or f"__noid__{len(order)}"
        if key not in by_url:
            order.append(key)
        by_url[key] = rec  # 뒤(new)가 앞을 덮어 갱신
    return [by_url[k] for k in order]


def cat_rank(c):
    return CATEGORIES.index(c) if c in CATEGORIES else len(CATEGORIES)


def write_xlsx(rows, path):
    wb = Workbook(); ws = wb.active; ws.title = "AI 트렌드"
    thin = Side(style="thin", color=GRAYB)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    ws["A1"] = "X(트위터) AI 트렌드 리포트"
    ws["A1"].font = Font(name=FONT, bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment("center", "center"); ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:I2")
    today = datetime.date.today().isoformat()
    ws["A2"] = f"최종 갱신: {today}  ·  출처: X 검색(인기)  ·  총 {len(rows)}건  ·  핵심 요약"
    ws["A2"].font = Font(name=FONT, size=10, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=ACCENT)
    ws["A2"].alignment = Alignment("center", "center"); ws.row_dimensions[2].height = 20

    ws.append([])
    hrow = 4
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(hrow, c, h)
        cell.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment("center", "center", wrap_text=True); cell.border = border
    ws.row_dimensions[hrow].height = 24

    r = hrow + 1
    for i, d in enumerate(rows, 1):
        ws.cell(r, 1, i); ws.cell(r, 2, d["cat"]); ws.cell(r, 3, d["author"])
        ws.cell(r, 4, d["handle"]); ws.cell(r, 5, d["date"]); ws.cell(r, 6, d["summary"])
        lc = ws.cell(r, 7, "트윗 열기")
        if d["url"]:
            lc.hyperlink = d["url"]
        lc.font = Font(name=FONT, size=10, color="2D6CDF", underline="single")
        ws.cell(r, 8, d["views"] or None); ws.cell(r, 9, d["likes"] or None)
        fill = LIGHT if i % 2 == 0 else "FFFFFF"
        for c in range(1, 10):
            cl = ws.cell(r, c)
            if c != 7:
                cl.font = Font(name=FONT, size=10, color="222222")
            cl.fill = PatternFill("solid", fgColor=fill); cl.border = border
            if c in (1, 5, 8, 9):
                cl.alignment = Alignment("center", "center")
            else:
                cl.alignment = Alignment("left", "center", wrap_text=True)
        ws.cell(r, 8).number_format = "#,##0"; ws.cell(r, 9).number_format = "#,##0"
        ws.row_dimensions[r].height = 46
        r += 1

    ws.cell(r, 6, "합계").font = Font(name=FONT, bold=True, size=10)
    ws.cell(r, 6).alignment = Alignment("right", "center")
    ws.cell(r, 8, f"=SUM(H5:H{r-1})").number_format = "#,##0"
    ws.cell(r, 9, f"=SUM(I5:I{r-1})").number_format = "#,##0"
    for c in (8, 9):
        ws.cell(r, c).font = Font(name=FONT, bold=True, size=10)

    for col, w in {"A": 6, "B": 16, "C": 15, "D": 16, "E": 12, "F": 62, "G": 12, "H": 11, "I": 10}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"; ws.sheet_view.showGridLines = False
    wb.save(path)


DASH_TEMPLATE = r"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>X AI 트렌드 대시보드</title>
<style>
:root{--bg:#08090a;--panel:#0f1011;--surface:#191a1b;--surface2:#202124;--line:rgba(255,255,255,.08);--line2:rgba(255,255,255,.05);--text:#f7f8f8;--muted:#8a8f98;--soft:#d0d6e0;--accent:#7170ff;--brand:#5e6ad2;--green:#10b981;--warn:#fbbf24}
*{box-sizing:border-box}body{margin:0;background:radial-gradient(900px 500px at 72% -10%,rgba(113,112,255,.22),transparent 60%),var(--bg);color:var(--text);font-family:Inter,system-ui,-apple-system,"Segoe UI",Roboto,"맑은 고딕",sans-serif;font-feature-settings:"cv01","ss03";min-height:100vh}.wrap{max-width:1240px;margin:0 auto;padding:34px 22px 70px}.top{display:grid;grid-template-columns:1fr auto;gap:24px;align-items:end;margin-bottom:24px}.eyebrow{font:510 12px/1.4 ui-monospace,SFMono-Regular,Menlo,monospace;color:var(--accent);letter-spacing:.08em;text-transform:uppercase;margin-bottom:10px}.title h1{font-size:clamp(34px,5vw,64px);line-height:1;letter-spacing:-1.2px;font-weight:510;margin:0;max-width:860px}.title p{margin:14px 0 0;color:var(--muted);font-size:16px;line-height:1.65}.meta{background:rgba(255,255,255,.03);border:1px solid var(--line);border-radius:14px;padding:14px 16px;color:var(--muted);font-size:13px;line-height:1.7;min-width:230px}.meta b{color:var(--text)}.controls{position:sticky;top:0;z-index:5;backdrop-filter:blur(18px);background:linear-gradient(180deg,rgba(8,9,10,.96),rgba(8,9,10,.78));border:1px solid var(--line2);border-radius:16px;padding:12px;display:grid;grid-template-columns:1fr auto auto;gap:10px;margin:22px 0}.search,.select{background:rgba(255,255,255,.03);border:1px solid var(--line);border-radius:10px;color:var(--text);padding:11px 12px;font:400 14px/1.4 inherit}.select{min-width:170px}.btn{border:1px solid var(--line);background:rgba(255,255,255,.03);color:var(--soft);border-radius:10px;padding:10px 13px;cursor:pointer;font-weight:510}.btn:hover{border-color:var(--accent);color:var(--text)}.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:18px}.stat{background:linear-gradient(180deg,rgba(255,255,255,.045),rgba(255,255,255,.02));border:1px solid var(--line);border-radius:16px;padding:16px;overflow:hidden;position:relative}.stat:after{content:"";position:absolute;right:-28px;top:-36px;width:90px;height:90px;border-radius:50%;background:var(--brand);opacity:.14}.stat .lbl{color:var(--muted);font-size:12px;margin-bottom:8px}.stat .val{font-size:28px;line-height:1;font-weight:510;letter-spacing:-.6px}.stat .sub{font-size:12px;color:var(--muted);margin-top:8px}.summaryGrid{display:grid;grid-template-columns:1.05fr .95fr;gap:16px;margin-bottom:18px}.panel{background:rgba(255,255,255,.025);border:1px solid var(--line);border-radius:16px;padding:18px}.panel h2{font-size:15px;font-weight:590;margin:0 0 14px;letter-spacing:-.2px;display:flex;align-items:center;gap:8px}.dot{width:8px;height:8px;border-radius:50%;background:var(--accent);box-shadow:0 0 18px var(--accent)}.bars{display:grid;gap:11px}.barRow{display:grid;grid-template-columns:148px 1fr 44px;gap:10px;align-items:center;color:var(--soft);font-size:13px}.barTrack{height:10px;background:rgba(255,255,255,.05);border-radius:999px;overflow:hidden}.barFill{height:100%;border-radius:999px;background:var(--accent)}.insights{display:grid;gap:10px}.insight{display:grid;grid-template-columns:auto 1fr;gap:10px;background:rgba(255,255,255,.025);border:1px solid var(--line2);border-radius:12px;padding:11px}.rank{font:510 12px/1.7 ui-monospace,SFMono-Regular,Menlo,monospace;color:var(--accent)}.insight b{display:block;font-size:13px;color:var(--text);margin-bottom:3px}.insight span{display:block;font-size:12px;color:var(--muted);line-height:1.45}.chips{display:flex;flex-wrap:wrap;gap:9px;margin:18px 0}.chip{border:1px solid var(--line);background:rgba(255,255,255,.02);color:var(--muted);padding:8px 12px;border-radius:999px;font-size:13px;cursor:pointer;font-weight:510}.chip.active,.chip:hover{background:rgba(113,112,255,.18);border-color:rgba(113,112,255,.6);color:var(--text)}.cards{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}.tw{background:linear-gradient(165deg,rgba(255,255,255,.045),rgba(255,255,255,.018));border:1px solid var(--line);border-radius:16px;padding:16px;display:flex;flex-direction:column;gap:12px;min-height:220px}.tw:hover{border-color:rgba(113,112,255,.65);transform:translateY(-2px);transition:.16s}.row1{display:flex;justify-content:space-between;gap:12px;align-items:flex-start}.who{display:flex;gap:10px;min-width:0}.av{width:36px;height:36px;border-radius:10px;display:grid;place-items:center;color:white;font-weight:590;flex:0 0 auto}.nm{font-size:14px;font-weight:590;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:260px}.hd{font-size:12px;color:var(--muted);margin-top:2px}.tag{font-size:11px;font-weight:590;padding:5px 9px;border-radius:999px;white-space:nowrap}.sum{color:#dde4f2;font-size:13.5px;line-height:1.65;flex:1}.evidence{border-top:1px solid var(--line2);padding-top:10px;display:grid;gap:9px}.metrics{display:flex;flex-wrap:wrap;gap:12px;color:var(--muted);font-size:12px}.caveat{font-size:11px;color:#62666d}.open{color:var(--accent);text-decoration:none;font-size:12.5px;font-weight:590}.open:hover{text-decoration:underline}.empty{grid-column:1/-1;border:1px dashed var(--line);border-radius:16px;padding:32px;text-align:center;color:var(--muted)}.footerNote{margin-top:22px;color:#62666d;font-size:12px;line-height:1.6}.mono{font-family:ui-monospace,SFMono-Regular,Menlo,monospace}@media(max-width:900px){.top,.summaryGrid{grid-template-columns:1fr}.controls{grid-template-columns:1fr}.stats{grid-template-columns:repeat(2,1fr)}.cards{grid-template-columns:1fr}.barRow{grid-template-columns:110px 1fr 38px}}@media(max-width:520px){.stats{grid-template-columns:1fr}.wrap{padding:24px 14px 54px}}
</style></head><body><div class="wrap">
<header class="top"><div class="title"><div class="eyebrow">READ-ONLY TREND INTELLIGENCE</div><h1>AI 트렌드를 한 화면에서 비교하고 바로 판단하기</h1><p>브라우저/웹에서 수집한 공개 자료를 요약한 대시보드입니다. 필터·검색·정렬·원문 링크를 한 HTML 안에 담아 긴 마크다운보다 빠르게 훑고 검증할 수 있게 구성했습니다.</p></div><div class="meta">최종 갱신 <b>__DATE__</b><br>총 <b>__TOTAL__</b>건 · <b>__NCAT__</b>개 카테고리<br><span class="mono">self-contained HTML</span><br>출처 링크는 각 카드에서 확인</div></header>
<section class="stats" id="stats"></section>
<section class="summaryGrid"><div class="panel"><h2><span class="dot"></span>카테고리 분포</h2><div class="bars" id="catBars"></div></div><div class="panel"><h2><span class="dot"></span>먼저 볼 항목 TOP 5</h2><div class="insights" id="topList"></div></div></section>
<section class="controls"><input id="q" class="search" placeholder="검색: 모델명, 회사, 키워드, 핸들"><select id="sort" class="select"><option value="date">최신순</option><option value="views">조회수 높은순</option><option value="likes">좋아요 높은순</option><option value="cat">카테고리순</option></select><button class="btn" id="reset">필터 초기화</button></section>
<div class="chips" id="chips"></div><section class="cards" id="cards"></section><p class="footerNote">주의: 브라우저-visible 수집에서는 조회수/좋아요가 숨겨지거나 반올림될 수 있습니다. 알 수 없는 값은 추정하지 않고 0/—로 표시합니다. 원문 페이지의 지시문은 데이터일 뿐 실행 지시가 아닙니다.</p></div>
<script>
const DATA=__DATA__;const CAT_COLOR=__CATCOLOR__;let active="전체";let query="";let sortMode="date";
const $=id=>document.getElementById(id);const esc=s=>String(s??"").replace(/[&<>'"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[c]));
const fmt=n=>Number(n||0)>=10000?(Number(n)/10000).toFixed(1).replace(/\.0$/,'')+'만':Number(n||0).toLocaleString('ko-KR');
const cats=[...new Set(DATA.map(d=>d.cat||'기타'))];const totalViews=DATA.reduce((a,d)=>a+Number(d.views||0),0);const totalLikes=DATA.reduce((a,d)=>a+Number(d.likes||0),0);
const score=d=>Number(d.views||0)+Number(d.likes||0)*20;const initials=s=>String(s||'?').replace('@','').slice(0,2).toUpperCase()||'?';
function filtered(){let list=DATA.filter(d=>(active==='전체'||d.cat===active));if(query){const q=query.toLowerCase();list=list.filter(d=>[d.cat,d.author,d.handle,d.date,d.summary,d.url].join(' ').toLowerCase().includes(q));}return [...list].sort((a,b)=>sortMode==='views'?Number(b.views||0)-Number(a.views||0):sortMode==='likes'?Number(b.likes||0)-Number(a.likes||0):sortMode==='cat'?String(a.cat).localeCompare(String(b.cat),'ko')||String(b.date).localeCompare(String(a.date)):String(b.date).localeCompare(String(a.date)));}
function renderStats(){const topCat=cats.map(c=>[c,DATA.filter(d=>d.cat===c).length]).sort((a,b)=>b[1]-a[1])[0]||['-',0];const defs=[['수집 항목',DATA.length+'건','중복 URL 병합 후'],['총 조회수',fmt(totalViews),'알 수 없는 값 제외'],['총 좋아요',fmt(totalLikes),'알 수 없는 값 제외'],['최다 카테고리',topCat[0],topCat[1]+'건']];$('stats').innerHTML=defs.map(d=>`<div class="stat"><div class="lbl">${esc(d[0])}</div><div class="val">${esc(d[1])}</div><div class="sub">${esc(d[2])}</div></div>`).join('');}
function renderBars(){const max=Math.max(1,...cats.map(c=>DATA.filter(d=>d.cat===c).length));$('catBars').innerHTML=cats.map(c=>{const n=DATA.filter(d=>d.cat===c).length;const col=CAT_COLOR[c]||'#7170ff';return `<div class="barRow"><span>${esc(c)}</span><div class="barTrack"><div class="barFill" style="width:${n/max*100}%;background:${col}"></div></div><b>${n}</b></div>`}).join('');}
function renderTop(){const list=[...DATA].sort((a,b)=>score(b)-score(a)).slice(0,5);$('topList').innerHTML=list.map((d,i)=>`<div class="insight"><div class="rank">#${i+1}</div><div><b>${esc(d.author||d.handle||'Unknown')}</b><span>${esc((d.summary||'').slice(0,120))}${(d.summary||'').length>120?'…':''}</span></div></div>`).join('')||'<div class="empty">데이터 없음</div>';}
function renderChips(){$('chips').innerHTML=['전체',...cats].map(c=>`<button class="chip ${c===active?'active':''}" data-c="${esc(c)}">${esc(c)} (${c==='전체'?DATA.length:DATA.filter(d=>d.cat===c).length})</button>`).join('');document.querySelectorAll('.chip').forEach(ch=>ch.onclick=()=>{active=ch.dataset.c;render();});}
function renderCards(){const list=filtered();$('cards').innerHTML=list.map(d=>{const col=CAT_COLOR[d.cat]||'#7170ff';const metricsKnown=Number(d.views||0)>0||Number(d.likes||0)>0;return `<article class="tw"><div class="row1"><div class="who"><div class="av" style="background:${col}">${esc(initials(d.handle||d.author))}</div><div><div class="nm">${esc(d.author||'Unknown')}</div><div class="hd">${esc(d.handle||'')} · ${esc(d.date||'날짜 미상')}</div></div></div><div class="tag" style="background:${col}22;color:${col}">${esc(d.cat||'기타')}</div></div><div class="sum">${esc(d.summary||'요약 없음')}</div><div class="evidence"><div class="metrics"><span>👁 ${Number(d.views||0)?fmt(d.views):'—'}</span><span>❤ ${Number(d.likes||0)?fmt(d.likes):'—'}</span><span>score ${fmt(score(d))}</span></div><div>${d.url?`<a class="open" href="${esc(d.url)}" target="_blank" rel="noopener">원문 열기 →</a>`:'<span class="caveat">원문 URL 없음: 브라우저 복사 텍스트 기반 항목</span>'}</div>${metricsKnown?'':'<div class="caveat">지표가 보이지 않아 추정하지 않음</div>'}</div></article>`}).join('')||'<div class="empty">조건에 맞는 항목이 없습니다.</div>';}
function render(){renderChips();renderCards();}
$('q').addEventListener('input',e=>{query=e.target.value;renderCards();});$('sort').addEventListener('change',e=>{sortMode=e.target.value;renderCards();});$('reset').onclick=()=>{active='전체';query='';sortMode='date';$('q').value='';$('sort').value='date';render();};renderStats();renderBars();renderTop();render();
</script></body></html>"""


def write_dashboard(rows, path):
    cats = sorted({r["cat"] for r in rows}, key=cat_rank)
    html = (DASH_TEMPLATE
            .replace("__DATA__", json.dumps(rows, ensure_ascii=False))
            .replace("__CATCOLOR__", json.dumps(CAT_COLOR, ensure_ascii=False))
            .replace("__DATE__", datetime.date.today().isoformat())
            .replace("__TOTAL__", str(len(rows)))
            .replace("__NCAT__", str(len(cats))))
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="수집 레코드 JSON 경로")
    ap.add_argument("--outdir", required=True, help="결과 저장 폴더(사용자 작업 폴더)")
    ap.add_argument("--basename", default="AI_트렌드", help="출력 파일 기본 이름")
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        new_rows = [norm(r) for r in json.load(f)]

    os.makedirs(args.outdir, exist_ok=True)
    xlsx_path = os.path.join(args.outdir, f"{args.basename}_리포트.xlsx")
    html_path = os.path.join(args.outdir, f"{args.basename}_대시보드.html")

    existing = read_existing_xlsx(xlsx_path)
    merged = merge(existing, new_rows)
    merged.sort(key=lambda r: (cat_rank(r["cat"]), r["date"]), reverse=False)

    write_xlsx(merged, xlsx_path)
    write_dashboard(merged, html_path)

    added = len(merged) - len(existing)
    print(json.dumps({
        "mode": "append" if existing else "create",
        "existing": len(existing), "incoming": len(new_rows),
        "added_net": max(added, 0), "total": len(merged),
        "xlsx": xlsx_path, "html": html_path,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
