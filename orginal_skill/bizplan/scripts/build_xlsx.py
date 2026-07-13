#!/usr/bin/env python3
"""
build_xlsx.py — YAML/JSON spec → XLSX 생성기 (openpyxl)

source-index · claim-chart · evidence-ledger · market-sizing · financial-model 등
표 산출물을 만든다.

사용:
  # spec 파일로 임의 시트 구성
  python3 build_xlsx.py <spec.yaml|spec.json> <out.xlsx>

  # spec 없이 표준 템플릿(헤더만) 생성
  python3 build_xlsx.py --preset source-index <out.xlsx>
  python3 build_xlsx.py --preset claim-chart <out.xlsx>
  python3 build_xlsx.py --preset evidence-ledger <out.xlsx>

-----------------------------------------------------------------------------
spec 형식 (YAML 예시):

  sheets:
    - name: "source-index"          # 시트 이름 (필수)
      freeze_header: true            # 첫 행 고정 (기본 true)
      columns:                       # 헤더 (문자열 리스트, 필수)
        - id
        - 제목
        - 발행기관
      widths: [8, 40, 20]            # (선택) 열 너비. 생략 시 내용 기반 자동
      rows:                          # (선택) 데이터 행 (리스트의 리스트)
        - ["S001", "2025 시장보고서", "통계청"]
        - ["S002", "특허 동향", "특허청"]

  여러 시트를 두려면 sheets 리스트에 항목을 추가한다.
  최상위가 dict 가 아니라 list 면 시트 리스트로 간주한다.
  숫자/날짜 셀은 타입을 보존하고, 긴 텍스트는 줄바꿈(wrap) 처리한다.
-----------------------------------------------------------------------------
"""
import argparse
import datetime
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[오류] openpyxl 이 없습니다.  pip3 install openpyxl", file=sys.stderr)
    raise SystemExit(2)

try:
    import yaml
    _HAS_YAML = True
except ImportError:
    _HAS_YAML = False


# 표준 프리셋 컬럼 (SKILL.md / references 규약과 일치)
PRESETS = {
    # references/04-research-engine.md §4 와 정확히 일치
    "source-index": [
        "id", "제목", "발행기관", "저자", "발행일", "기준시점", "접근일",
        "URL", "사용가능주장", "신뢰도", "반영위치", "stance", "priority_rank",
    ],
    # SKILL.md §123 + 작업 명세
    "claim-chart": [
        "청구항요소", "사용자제품", "일치여부", "비고", "위험도",
    ],
    # 작업 명세: evidence-ledger
    "evidence-ledger": [
        "claim_text", "claim_type", "source", "confidence",
        "target_section", "verification_status",
    ],
}

# 프리셋별 안내 행(데이터 입력 가이드용 주석) — 신뢰도/태그 허용값 등
PRESET_NOTES = {
    "source-index": "신뢰도=상/중/하 · stance=우호/중립/부정·반대 · priority_rank=1~10",
    "claim-chart": "일치여부=일치/부분일치/불일치 · 위험도=상/중/하 (변리사 검토 대상 표시)",
    "evidence-ledger": "claim_type=사실/추정/가정/목표/확인필요 · confidence=상/중/하 · verification_status=확인됨/미확인",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _load_spec(path):
    with open(path, encoding="utf-8") as f:
        text = f.read()
    if path.lower().endswith(".json"):
        return json.loads(text)
    if not _HAS_YAML:
        print("[오류] YAML spec 인데 PyYAML 이 없습니다.  pip3 install pyyaml", file=sys.stderr)
        raise SystemExit(2)
    return yaml.safe_load(text)


def _normalize_spec(spec):
    """다양한 형태의 spec 을 [{name, columns, rows, widths, freeze_header}] 로 정규화."""
    if spec is None:
        return []
    if isinstance(spec, list):
        sheets = spec
    elif isinstance(spec, dict) and "sheets" in spec:
        sheets = spec["sheets"]
    elif isinstance(spec, dict):
        # 단일 시트 dict 로 간주
        sheets = [spec]
    else:
        raise ValueError("알 수 없는 spec 형식")
    out = []
    for s in sheets:
        if not isinstance(s, dict):
            continue
        out.append({
            "name": str(s.get("name", "Sheet")),
            "columns": list(s.get("columns", [])),
            "rows": list(s.get("rows", [])),
            "widths": list(s.get("widths", [])),
            "freeze_header": bool(s.get("freeze_header", True)),
        })
    return out


def _coerce(value):
    """문자열이 숫자/날짜로 보이면 해당 타입으로 변환(셀 타입 보존)."""
    if not isinstance(value, str):
        return value
    v = value.strip()
    if v == "":
        return None
    # 천단위 콤마 제거 후 숫자 판정 (단, 단위/기호가 섞이면 문자열 유지)
    cleaned = v.replace(",", "")
    body = cleaned.lstrip("-")
    # 정수 (콤마 천단위 허용: 1,200,000 → 1200000)
    try:
        if body.isdigit():
            return int(cleaned)
    except Exception:
        pass
    # 실수
    try:
        if "." in body and body.replace(".", "", 1).isdigit():
            return float(cleaned)
    except Exception:
        pass
    # 날짜 YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return value


def _write_sheet(ws, sheet):
    columns = sheet["columns"]
    rows = sheet["rows"]
    ncols = len(columns)

    # 헤더
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # 데이터
    for r, row in enumerate(rows, start=2):
        for c in range(ncols):
            raw = row[c] if c < len(row) else None
            value = _coerce(raw)
            cell = ws.cell(row=r, column=c + 1, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 열 너비
    widths = sheet["widths"]
    for c in range(ncols):
        letter = get_column_letter(c + 1)
        if c < len(widths) and widths[c]:
            ws.column_dimensions[letter].width = float(widths[c])
        else:
            # 내용 기반 자동(헤더 + 데이터 최대 길이, 한글 가중)
            maxlen = len(str(columns[c]))
            for row in rows:
                if c < len(row) and row[c] is not None:
                    maxlen = max(maxlen, len(str(row[c])))
            # 한글 고려해 약간 넉넉히, 상한 60
            ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 60)

    # 헤더 행 높이
    ws.row_dimensions[1].height = 22

    # 필터 + 고정
    if ncols > 0:
        last_col = get_column_letter(ncols)
        last_row = max(1, len(rows) + 1)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    if sheet["freeze_header"]:
        ws.freeze_panes = "A2"

    # 캡쳐(xlsx_to_image.sh) 품질용 인쇄설정 — 표가 한 페이지 너비에 맞게 (2026-06-18)
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 너비만 맞춤(세로로 길면 여러 페이지 허용)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    for _m in ("left", "right", "top", "bottom"):
        setattr(ws.page_margins, _m, 0.3)


def main() -> int:
    ap = argparse.ArgumentParser(description="YAML/JSON spec → XLSX (openpyxl)")
    ap.add_argument("spec", nargs="?", help="spec 파일 (.yaml/.json). --preset 사용 시 생략")
    ap.add_argument("output", nargs="?", help="출력 .xlsx 파일")
    ap.add_argument("--preset", choices=sorted(PRESETS.keys()),
                    help="spec 없이 표준 템플릿(헤더만) 생성")
    args = ap.parse_args()

    # --preset 모드: 인자가 (output) 1개로 들어올 수 있음
    if args.preset:
        # preset 모드에서는 첫 위치인자가 output
        out = args.output or args.spec
        if not out:
            ap.print_usage()
            print("\n예) python3 build_xlsx.py --preset source-index source-index.xlsx", file=sys.stderr)
            return 1
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = args.preset
        sheet = {
            "name": args.preset,
            "columns": PRESETS[args.preset],
            "rows": [],
            "widths": [],
            "freeze_header": True,
        }
        _write_sheet(ws, sheet)
        # 안내 메모를 헤더 셀 주석으로
        note = PRESET_NOTES.get(args.preset)
        if note:
            try:
                from openpyxl.comments import Comment
                ws.cell(row=1, column=1).comment = Comment(note, "bizplan")
            except Exception:
                pass
        _save(wb, out)
        print(f"[완료] 프리셋 XLSX 생성: {out}  (preset={args.preset}, 헤더 {len(PRESETS[args.preset])}컬럼)")
        if note:
            print(f"  입력 가이드: {note}")
        return 0

    # spec 모드
    if not args.spec or not args.output:
        ap.print_usage()
        print("\n예) python3 build_xlsx.py spec.yaml out.xlsx", file=sys.stderr)
        print("    python3 build_xlsx.py --preset source-index source-index.xlsx", file=sys.stderr)
        return 1

    if not os.path.exists(args.spec):
        print(f"[오류] spec 파일 없음: {args.spec}", file=sys.stderr)
        return 1

    try:
        raw = _load_spec(args.spec)
        sheets = _normalize_spec(raw)
    except Exception as e:
        print(f"[오류] spec 파싱 실패: {e}", file=sys.stderr)
        return 1

    if not sheets:
        print("[오류] spec 에 시트가 없습니다 (sheets: 또는 columns: 필요)", file=sys.stderr)
        return 1

    wb = openpyxl.Workbook()
    # 기본 시트 제거 후 spec 대로 추가
    default = wb.active
    wb.remove(default)
    for sheet in sheets:
        title = sheet["name"][:31] or "Sheet"  # 엑셀 시트명 31자 제한
        ws = wb.create_sheet(title=title)
        if not sheet["columns"]:
            print(f"[경고] 시트 '{title}' 에 columns 가 없습니다 — 빈 시트로 생성", file=sys.stderr)
            continue
        _write_sheet(ws, sheet)

    _save(wb, args.output)
    total_rows = sum(len(s["rows"]) for s in sheets)
    print(f"[완료] XLSX 생성: {args.output}  (시트 {len(sheets)}개, 데이터 {total_rows}행)")
    return 0


def _save(wb, out):
    out_dir = os.path.dirname(os.path.abspath(out))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out)


if __name__ == "__main__":
    raise SystemExit(main())
